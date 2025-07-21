import pandas as pd
import openpyxl
import logging
from typing import List, Any, Optional, Union
from pathlib import Path


class ExcelReader:
    """Utility class for reading Excel files"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.logger = logging.getLogger(self.__class__.__name__)
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    def read_sheet(self, sheet_name: str = None, header_row: int = 0) -> pd.DataFrame:
        """
        Read a specific sheet from Excel file
        
        Args:
            sheet_name (str): Name of the sheet to read. If None, reads first sheet
            header_row (int): Row number to use as header (0-indexed)
        
        Returns:
            pd.DataFrame: DataFrame containing the sheet data
        """
        try:
            if sheet_name:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
            else:
                df = pd.read_excel(self.file_path, header=header_row)
            
            self.logger.info(f"Successfully read sheet '{sheet_name}' from {self.file_path}")
            return df
        
        except Exception as e:
            self.logger.error(f"Error reading Excel sheet: {e}")
            raise
    
    def read_all_sheets(self) -> dict[str, pd.DataFrame]:
        """
        Read all sheets from Excel file
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            all_sheets = pd.read_excel(self.file_path, sheet_name=None)
            self.logger.info(f"Successfully read all sheets from {self.file_path}")
            return all_sheets
        
        except Exception as e:
            self.logger.error(f"Error reading all Excel sheets: {e}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of all sheet names in the Excel file
        
        Returns:
            List[str]: List of sheet names
        """
        try:
            with pd.ExcelFile(self.file_path) as xls:
                sheet_names = xls.sheet_names
            
            self.logger.info(f"Found {len(sheet_names)} sheets in {self.file_path}")
            return sheet_names
        
        except Exception as e:
            self.logger.error(f"Error getting sheet names: {e}")
            raise
    
    def read_cell(self, sheet_name: str, row: int, column: Union[int, str]) -> Any:
        """
        Read a specific cell value
        
        Args:
            sheet_name (str): Name of the sheet
            row (int): Row number (1-indexed)
            column (Union[int, str]): Column number (1-indexed) or column letter
        
        Returns:
            Any: Cell value
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[sheet_name]
            
            if isinstance(column, str):
                cell_value = worksheet[f"{column}{row}"].value
            else:
                cell_value = worksheet.cell(row=row, column=column).value
            
            self.logger.info(f"Read cell value from {sheet_name}[{row}, {column}]: {cell_value}")
            return cell_value
        
        except Exception as e:
            self.logger.error(f"Error reading cell: {e}")
            raise
    
    def read_range(self, sheet_name: str, start_cell: str, end_cell: str) -> List[List[Any]]:
        """
        Read a range of cells
        
        Args:
            sheet_name (str): Name of the sheet
            start_cell (str): Starting cell (e.g., 'A1')
            end_cell (str): Ending cell (e.g., 'C3')
        
        Returns:
            List[List[Any]]: 2D list containing the range values
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[sheet_name]
            
            cell_range = worksheet[f"{start_cell}:{end_cell}"]
            values = []
            
            for row in cell_range:
                row_values = [cell.value for cell in row]
                values.append(row_values)
            
            self.logger.info(f"Read range {start_cell}:{end_cell} from {sheet_name}")
            return values
        
        except Exception as e:
            self.logger.error(f"Error reading range: {e}")
            raise
    
    def read_column(self, sheet_name: str, column: Union[int, str], skip_header: bool = True) -> List[Any]:
        """
        Read entire column values
        
        Args:
            sheet_name (str): Name of the sheet
            column (Union[int, str]): Column number (1-indexed) or column letter
            skip_header (bool): Whether to skip the first row (header)
        
        Returns:
            List[Any]: List of column values
        """
        try:
            df = self.read_sheet(sheet_name)
            
            if isinstance(column, str):
                col_values = df[column].tolist()
            else:
                col_values = df.iloc[:, column-1].tolist()
            
            if skip_header:
                col_values = col_values[1:]
            
            self.logger.info(f"Read column {column} from {sheet_name}")
            return col_values
        
        except Exception as e:
            self.logger.error(f"Error reading column: {e}")
            raise
    
    def read_row(self, sheet_name: str, row_number: int, skip_header_col: bool = False) -> List[Any]:
        """
        Read entire row values
        
        Args:
            sheet_name (str): Name of the sheet
            row_number (int): Row number (1-indexed)
            skip_header_col (bool): Whether to skip the first column
        
        Returns:
            List[Any]: List of row values
        """
        try:
            df = self.read_sheet(sheet_name)
            row_values = df.iloc[row_number-1].tolist()
            
            if skip_header_col:
                row_values = row_values[1:]
            
            self.logger.info(f"Read row {row_number} from {sheet_name}")
            return row_values
        
        except Exception as e:
            self.logger.error(f"Error reading row: {e}")
            raise
    
    def get_row_count(self, sheet_name: str = None) -> int:
        """
        Get number of rows in sheet
        
        Args:
            sheet_name (str): Name of the sheet
        
        Returns:
            int: Number of rows
        """
        try:
            df = self.read_sheet(sheet_name)
            row_count = len(df)
            
            self.logger.info(f"Sheet '{sheet_name}' has {row_count} rows")
            return row_count
        
        except Exception as e:
            self.logger.error(f"Error getting row count: {e}")
            raise
    
    def get_column_count(self, sheet_name: str = None) -> int:
        """
        Get number of columns in sheet
        
        Args:
            sheet_name (str): Name of the sheet
        
        Returns:
            int: Number of columns
        """
        try:
            df = self.read_sheet(sheet_name)
            col_count = len(df.columns)
            
            self.logger.info(f"Sheet '{sheet_name}' has {col_count} columns")
            return col_count
        
        except Exception as e:
            self.logger.error(f"Error getting column count: {e}")
            raise
    
    def search_value(self, sheet_name: str, search_value: Any) -> List[tuple]:
        """
        Search for a value in the sheet and return its positions
        
        Args:
            sheet_name (str): Name of the sheet
            search_value (Any): Value to search for
        
        Returns:
            List[tuple]: List of (row, column) positions where value was found
        """
        try:
            df = self.read_sheet(sheet_name)
            positions = []
            
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    if value == search_value:
                        positions.append((row_idx + 1, col_idx + 1))  # Convert to 1-indexed
            
            self.logger.info(f"Found '{search_value}' at {len(positions)} positions in {sheet_name}")
            return positions
        
        except Exception as e:
            self.logger.error(f"Error searching value: {e}")
            raise
